"""
ROIN WORLD FZE — Monthly Schedule Plan Handler
================================================
Section B of SCHEDULE_FEEDBACK_PROMPT.md

Sheet tab: Monthly_Schedule
Columns: Plan_ID, Month, Department, Location, Emp_Code, Emp_Name(VLOOKUP),
         Shift, D1-D31, Uploaded_By, Status, Created_At

Status flow: Draft → HR_Review → Approved → Active

Day values: W=Work, N=Night, OFF=Day Off, V=Vacation,
            S=Sick, U=Unpaid, B=Business Trip, H=Holiday, R=Rest
"""

import io
import os
from datetime import datetime, date
from calendar import monthrange

from telegram import InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (ConversationHandler, CallbackQueryHandler,
                           MessageHandler, filters)
from config import get_sheet

try:
    import openpyxl
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

try:
    from fpdf import FPDF
    _HAS_FPDF = True
except ImportError:
    _HAS_FPDF = False

# ── Helpers ──────────────────────────────────────────────────────────────────
def _bm():   return InlineKeyboardButton("↩️ Main Menu",     callback_data="back_to_menu")
def _bsch(): return InlineKeyboardButton("↩️ Schedule Menu", callback_data="menu_schedule")

# States
SCH_UPLOAD         = 2000
SCH_UPLOAD_CONFIRM = 2001
SCH_EDIT_DAY       = 2010
SCH_EDIT_VAL       = 2011
SCH_REVIEW_ACTION  = 2020

WORK_LOCATIONS = [
    "Office", "Russian Kitchen", "Egyptian Kitchen",
    "Pastry-Soup Kitchen", "FoodTruck & Bakery"
]

DAY_VALUES = ["W", "N", "OFF", "V", "S", "U", "B", "H", "R"]
DAY_LABELS = {
    "W": "Work (Day)",  "N": "Night Shift", "OFF": "Day Off",
    "V": "Vacation",    "S": "Sick Leave",  "U": "Unpaid Leave",
    "B": "Business",    "H": "Holiday",     "R": "Rest",
}
STATUS_FLOW = ["Draft", "HR_Review", "Approved", "Active"]

MGMT_ROLES = {"Bot_Manager", "Director", "HR_Manager", "HR_Staff",
              "Direct_Manager", "Supervisor"}
EDIT_ROLES  = {"Bot_Manager", "HR_Manager", "Direct_Manager"}
REVIEW_ROLES = {"Bot_Manager", "Director", "HR_Manager", "HR_Staff"}
APPROVE_ROLES = {"Bot_Manager", "Director", "HR_Manager"}


def _get_user(tid):
    """Return (emp_code, role, dept) for a Telegram user ID."""
    for i, r in enumerate(get_sheet("User_Registry").get_all_values()):
        if i == 0: continue
        if r[1].strip() == str(tid):
            ec   = r[0].strip()
            role = r[3].strip() if len(r) > 3 else "Employee"
            dept = ""
            for j, e in enumerate(get_sheet("Employee_DB").get_all_values()):
                if j == 0: continue
                if e[0].strip() == ec:
                    dept = e[6].strip() if len(e) > 6 else ""
                    break
            return ec, role, dept
    return None, "Employee", ""


def _gen_plan_id():
    ids = get_sheet("Monthly_Schedule").col_values(1)
    yr  = datetime.now().strftime("%Y")
    px  = f"SCH-{yr}-"
    mx  = 0
    for v in ids:
        if str(v).startswith(px):
            try: mx = max(mx, int(str(v).split("-")[-1]))
            except: pass
    return f"{px}{mx+1:04d}"


def _current_month_str():
    return datetime.now().strftime("%m/%Y")


def _month_days(month_str):
    """Return number of days in the given MM/YYYY month."""
    try:
        m, y = int(month_str[:2]), int(month_str[3:])
        return monthrange(y, m)[1]
    except Exception:
        return 31


def _parse_xlsx(file_bytes, dept, uploader_ec):
    """Parse uploaded xlsx, return list of row dicts or raise ValueError."""
    if not _HAS_OPENPYXL:
        raise ValueError("openpyxl not installed.")
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    headers = [str(c.value or "").strip() for c in ws[1]]
    rows = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        r = {headers[i]: str(raw[i] or "").strip()
             for i in range(min(len(headers), len(raw)))}
        ec = r.get("Emp_Code","").strip()
        if not ec: continue
        shift = r.get("Shift","Day")
        day_vals = {}
        for col in headers:
            if col.startswith("D") and col[1:].isdigit():
                v = r.get(col,"").upper().strip()
                if v not in DAY_VALUES: v = "W"
                day_vals[col] = v
        rows.append({"ec": ec, "shift": shift, "days": day_vals})
    if not rows:
        raise ValueError("No employee rows found in file.")
    return rows


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN SCHEDULE MENU
# ══════════════════════════════════════════════════════════════════════════════
async def sched_menu_handler(update, context):
    q = update.callback_query; await q.answer()
    ec, role, dept = _get_user(str(q.from_user.id))
    if role not in MGMT_ROLES and role != "Bot_Manager":
        await q.edit_message_text("❌ Access denied.",
                                  reply_markup=InlineKeyboardMarkup([[_bm()]])); return
    kb = []
    if role in EDIT_ROLES:
        kb.append([InlineKeyboardButton("📤 Upload Plan (Excel)",  callback_data="sched_upload")])
    kb.append([InlineKeyboardButton("📋 View Current Month",   callback_data="sched_view")])
    kb.append([InlineKeyboardButton("📅 Daily View",           callback_data="sched_daily")])
    kb.append([InlineKeyboardButton("📊 Plan vs Actual",       callback_data="sched_vs_actual")])
    if role in REVIEW_ROLES:
        kb.append([InlineKeyboardButton("✅ Review Plans",     callback_data="sched_review_list")])
    kb.append([InlineKeyboardButton("📄 Download PDF",         callback_data="sched_pdf")])
    kb.append([InlineKeyboardButton("↩️ Back", callback_data="back_to_menu"), _bm()])
    await q.edit_message_text(
        f"📅 Monthly Schedule\n\nDepartment: {dept or 'All'}\nMonth: {_current_month_str()}",
        reply_markup=InlineKeyboardMarkup(kb))


# ══════════════════════════════════════════════════════════════════════════════
#  UPLOAD PLAN
# ══════════════════════════════════════════════════════════════════════════════
async def sched_upload_start(update, context):
    q = update.callback_query; await q.answer()
    ec, role, dept = _get_user(str(q.from_user.id))
    if role not in EDIT_ROLES:
        await q.edit_message_text("❌ Access denied.", reply_markup=InlineKeyboardMarkup([[_bm()]])); return
    context.user_data["sched_dept"]     = dept
    context.user_data["sched_uploader"] = ec
    await q.edit_message_text(
        "📤 Upload Monthly Schedule Plan\n\n"
        "Send an Excel (.xlsx) file with columns:\n"
        "  Emp_Code | Shift | D1 | D2 | ... | D31\n\n"
        "Day values:\n"
        "  W=Work  N=Night  OFF=Day Off\n"
        "  V=Vacation  S=Sick  U=Unpaid\n"
        "  B=Business  H=Holiday  R=Rest\n\n"
        "Send the file now:",
        reply_markup=InlineKeyboardMarkup([[_bsch(), _bm()]]))
    return SCH_UPLOAD


async def sched_upload_doc(update, context):
    doc = update.message.document
    if not doc or not (doc.file_name or "").lower().endswith(".xlsx"):
        await update.message.reply_text(
            "⚠️ Please send a .xlsx Excel file.",
            reply_markup=InlineKeyboardMarkup([[_bsch(), _bm()]]))
        return SCH_UPLOAD
    await update.message.reply_text("⏳ Parsing file...")
    try:
        f_obj  = await doc.get_file()
        raw    = await f_obj.download_as_bytearray()
        dept   = context.user_data.get("sched_dept","")
        ec_up  = context.user_data.get("sched_uploader","")
        month  = _current_month_str()
        rows   = _parse_xlsx(bytes(raw), dept, ec_up)
        context.user_data["sched_rows"]  = rows
        context.user_data["sched_month"] = month
        n_days = _month_days(month)
        summary = (f"✅ File parsed!\n{'─'*24}\n"
                   f"Employees: {len(rows)}\n"
                   f"Month: {month}\n"
                   f"Department: {dept or 'Unknown'}\n"
                   f"Days: {n_days}\n\n"
                   "Save this plan as Draft?")
        kb = [[InlineKeyboardButton("✅ Save Draft", callback_data="sched_up_confirm"),
               InlineKeyboardButton("❌ Cancel",    callback_data="sched_up_cancel")],
              [_bm()]]
        await update.message.reply_text(summary, reply_markup=InlineKeyboardMarkup(kb))
        return SCH_UPLOAD_CONFIRM
    except Exception as e:
        await update.message.reply_text(
            f"❌ Parse error: {e}\n\nPlease check the file format and try again.",
            reply_markup=InlineKeyboardMarkup([[_bsch(), _bm()]]))
        return SCH_UPLOAD


async def sched_upload_confirm_cb(update, context):
    q = update.callback_query; await q.answer()
    if q.data == "sched_up_cancel":
        await q.edit_message_text("Cancelled.", reply_markup=InlineKeyboardMarkup([[_bsch(), _bm()]]))
        return ConversationHandler.END
    rows   = context.user_data.get("sched_rows", [])
    month  = context.user_data.get("sched_month", _current_month_str())
    dept   = context.user_data.get("sched_dept","")
    ec_up  = context.user_data.get("sched_uploader","")
    now    = datetime.now().strftime("%d/%m/%Y %H:%M")
    ws     = get_sheet("Monthly_Schedule")
    saved  = 0
    try:
        for r in rows:
            pid   = _gen_plan_id()
            ec    = r["ec"]
            shift = r.get("shift","Day")
            days  = r.get("days",{})
            # Cols: Plan_ID, Month, Department, Location, Emp_Code, Emp_Name(VLOOKUP),
            #       Shift, D1-D31(31 cols), Uploaded_By, Status, Created_At
            day_vals = [days.get(f"D{i}","") for i in range(1, 32)]
            row = [pid, month, dept, "", ec, "",
                   shift] + day_vals + [ec_up, "Draft", now]
            ws.append_row(row, value_input_option="USER_ENTERED")
            saved += 1
        await q.edit_message_text(
            f"✅ Schedule plan saved!\n{saved} employees uploaded.\n"
            f"Status: Draft\nMonth: {month}\n\n"
            "HR can now review this plan.",
            reply_markup=InlineKeyboardMarkup([[_bsch(), _bm()]]))
    except Exception as e:
        await q.edit_message_text(f"❌ Save error: {e}",
                                  reply_markup=InlineKeyboardMarkup([[_bm()]]))
    return ConversationHandler.END


async def sched_upload_cancel(update, context):
    await update.message.reply_text("Cancelled.")
    return ConversationHandler.END


# ══════════════════════════════════════════════════════════════════════════════
#  VIEW CURRENT MONTH PLAN
# ══════════════════════════════════════════════════════════════════════════════
async def sched_view_handler(update, context):
    q = update.callback_query; await q.answer()
    ec, role, dept = _get_user(str(q.from_user.id))
    await q.edit_message_text("⏳ Loading schedule...")
    try:
        all_rows = get_sheet("Monthly_Schedule").get_all_records()
    except Exception as e:
        await q.edit_message_text(f"❌ {e}", reply_markup=InlineKeyboardMarkup([[_bm()]])); return
    month = _current_month_str()
    if role in APPROVE_ROLES:
        rows = [r for r in all_rows if str(r.get("Month","")).strip() == month]
    else:
        rows = [r for r in all_rows
                if str(r.get("Month","")).strip() == month
                and str(r.get("Department","")).strip() == dept]
    if not rows:
        await q.edit_message_text(
            f"📋 No schedule for {month}.\n\nUpload a plan to get started.",
            reply_markup=InlineKeyboardMarkup([[_bsch(), _bm()]])); return
    # Show this week's columns (today's day)
    today_day = datetime.now().day
    start_d = max(1, today_day - 3)
    end_d   = min(31, today_day + 3)
    lines = [f"📋 Schedule — {month}\n{'─'*28}"]
    for r in rows[:25]:
        emp  = str(r.get("Emp_Code",""))
        dept_r = str(r.get("Department",""))
        week_vals = []
        for d in range(start_d, end_d + 1):
            v = str(r.get(f"D{d}","")).strip() or "?"
            week_vals.append(v)
        today_val = str(r.get(f"D{today_day}","")).strip() or "?"
        lines.append(f"• {emp} [{dept_r}] — Today: {today_val} | "
                     f"D{start_d}-D{end_d}: {' '.join(week_vals)}")
    lines.append(f"\n{'─'*28}\nTotal: {len(rows)} employee records")
    await q.edit_message_text("\n".join(lines),
                              reply_markup=InlineKeyboardMarkup([[_bsch(), _bm()]]))


# ══════════════════════════════════════════════════════════════════════════════
#  DAILY VIEW
# ══════════════════════════════════════════════════════════════════════════════
async def sched_daily_handler(update, context):
    q = update.callback_query; await q.answer()
    ec, role, dept = _get_user(str(q.from_user.id))
    await q.edit_message_text("⏳ Loading daily view...")
    try:
        all_rows = get_sheet("Monthly_Schedule").get_all_records()
    except Exception as e:
        await q.edit_message_text(f"❌ {e}", reply_markup=InlineKeyboardMarkup([[_bm()]])); return
    month   = _current_month_str()
    day_col = f"D{datetime.now().day}"
    today   = datetime.now().strftime("%d/%m/%Y")
    if role in APPROVE_ROLES:
        rows = [r for r in all_rows if str(r.get("Month","")).strip() == month]
    else:
        rows = [r for r in all_rows
                if str(r.get("Month","")).strip() == month
                and str(r.get("Department","")).strip() == dept]
    if not rows:
        await q.edit_message_text(
            f"📅 No schedule for {month}.",
            reply_markup=InlineKeyboardMarkup([[_bsch(), _bm()]])); return
    # Group by location and shift
    by_loc: dict = {}
    for r in rows:
        val   = str(r.get(day_col,"")).strip()
        loc   = str(r.get("Location","")).strip() or "Unknown"
        shift = str(r.get("Shift","Day")).strip()
        key   = (loc, shift)
        by_loc.setdefault(key, {"W": [], "N": [], "off": []})
        if val == "W":  by_loc[key]["W"].append(str(r.get("Emp_Code","")))
        elif val == "N": by_loc[key]["N"].append(str(r.get("Emp_Code","")))
        else:            by_loc[key]["off"].append(f"{r.get('Emp_Code','')}({val})")
    lines = [f"📅 Today — {dept or 'All Depts'}\n📆 {today}\n{'─'*28}"]
    for (loc, shift), counts in sorted(by_loc.items()):
        lines.append(f"\n📍 {loc} — {shift} Shift")
        lines.append(f"  ✅ Working: {len(counts['W'])} | "
                     f"🌙 Night: {len(counts['N'])}")
        if counts["off"]:
            lines.append(f"  ⬜ Off/Leave: {len(counts['off'])}")
    total_work = sum(len(v["W"]) + len(v["N"]) for v in by_loc.values())
    lines.append(f"\n{'─'*28}\nTotal scheduled today: {total_work}")
    await q.edit_message_text("\n".join(lines),
                              reply_markup=InlineKeyboardMarkup([[_bsch(), _bm()]]))


# ══════════════════════════════════════════════════════════════════════════════
#  PLAN VS ACTUAL
# ══════════════════════════════════════════════════════════════════════════════
async def sched_vs_actual_handler(update, context):
    q = update.callback_query; await q.answer()
    ec, role, dept = _get_user(str(q.from_user.id))
    await q.edit_message_text("⏳ Comparing plan vs actual...")
    try:
        sched_rows  = get_sheet("Monthly_Schedule").get_all_records()
        actual_rows = get_sheet("Attendance_Actual").get_all_records()
    except Exception as e:
        await q.edit_message_text(f"❌ {e}", reply_markup=InlineKeyboardMarkup([[_bm()]])); return
    month   = _current_month_str()
    day_col = f"D{datetime.now().day}"
    today   = datetime.now().strftime("%d/%m/%Y")
    if role in APPROVE_ROLES:
        plan = [r for r in sched_rows if str(r.get("Month","")).strip() == month]
    else:
        plan = [r for r in sched_rows
                if str(r.get("Month","")).strip() == month
                and str(r.get("Department","")).strip() == dept]
    # Employees planned to work today
    planned_work = [r for r in plan
                    if str(r.get(day_col,"")).strip() in ("W", "N")]
    planned_codes = {str(r.get("Emp_Code","")).strip() for r in planned_work}
    # Actual attendance today
    actual_codes = set()
    for r in actual_rows:
        if str(r.get("Date","")).strip() == today:
            actual_codes.add(str(r.get("Emp_Code","")).strip())
    present = planned_codes & actual_codes
    absent  = planned_codes - actual_codes
    extra   = actual_codes - planned_codes
    lines = [f"📊 Plan vs Actual — {today}\n{'─'*28}",
             f"✅ Planned to work: {len(planned_codes)}",
             f"✅ Present:         {len(present)}",
             f"❌ Absent (gap):    {len(absent)}",
             f"➕ Unplanned in:   {len(extra)}",
             f"{'─'*28}"]
    if absent:
        lines.append("❌ Absent employees:")
        for c in sorted(absent)[:15]:
            lines.append(f"  • {c}")
    await q.edit_message_text("\n".join(lines),
                              reply_markup=InlineKeyboardMarkup([[_bsch(), _bm()]]))


# ══════════════════════════════════════════════════════════════════════════════
#  REVIEW & APPROVAL
# ══════════════════════════════════════════════════════════════════════════════
async def sched_review_list_handler(update, context):
    q = update.callback_query; await q.answer()
    ec, role, dept = _get_user(str(q.from_user.id))
    if role not in REVIEW_ROLES:
        await q.edit_message_text("❌ Access denied.", reply_markup=InlineKeyboardMarkup([[_bm()]])); return
    await q.edit_message_text("⏳ Loading plans...")
    try:
        all_rows = get_sheet("Monthly_Schedule").get_all_records()
    except Exception as e:
        await q.edit_message_text(f"❌ {e}", reply_markup=InlineKeyboardMarkup([[_bm()]])); return
    # Group by Plan_ID+Month+Dept+Status — show unique plan summaries
    seen = {}
    for r in all_rows:
        key = (str(r.get("Month","")), str(r.get("Department","")))
        st  = str(r.get("Status","")).strip()
        if key not in seen:
            seen[key] = {"month": r.get("Month",""), "dept": r.get("Department",""),
                         "status": st, "count": 0}
        seen[key]["count"] += 1
    plans = list(seen.values())
    pending = [p for p in plans if p["status"] in ("Draft","HR_Review")]
    if not pending:
        await q.edit_message_text("✅ No plans pending review.",
                                  reply_markup=InlineKeyboardMarkup([[_bsch(), _bm()]])); return
    kb = []
    for p in pending[:15]:
        cb = f"sched_review_{p['month'].replace('/','_')}_{p['dept'][:10]}"
        kb.append([InlineKeyboardButton(
            f"📋 {p['dept']} — {p['month']} [{p['status']}] ({p['count']} emps)",
            callback_data=cb)])
    kb.append([_bsch(), _bm()])
    await q.edit_message_text(
        f"✅ Plans Pending Review ({len(pending)})\n\nSelect a plan:",
        reply_markup=InlineKeyboardMarkup(kb))


async def sched_review_cb(update, context):
    q = update.callback_query; await q.answer()
    ec, role, dept_u = _get_user(str(q.from_user.id))
    if role not in REVIEW_ROLES:
        await q.edit_message_text("❌ Access denied.", reply_markup=InlineKeyboardMarkup([[_bm()]])); return
    # parse month and dept from callback
    parts = q.data.replace("sched_review_","").split("_")
    month = f"{parts[0]}/{parts[1]}" if len(parts) >= 2 else ""
    dept  = "_".join(parts[2:]) if len(parts) > 2 else ""
    context.user_data["review_month"] = month
    context.user_data["review_dept"]  = dept
    try:
        rows = get_sheet("Monthly_Schedule").get_all_records()
    except Exception as e:
        await q.edit_message_text(f"❌ {e}", reply_markup=InlineKeyboardMarkup([[_bm()]])); return
    plan_rows = [r for r in rows
                 if str(r.get("Month","")).strip() == month
                 and str(r.get("Department","")).strip().startswith(dept)]
    if not plan_rows:
        await q.edit_message_text("❌ Plan not found.", reply_markup=InlineKeyboardMarkup([[_bm()]])); return
    status = str(plan_rows[0].get("Status","")).strip()
    day_today = datetime.now().day
    sample_lines = []
    for r in plan_rows[:5]:
        today_val = str(r.get(f"D{day_today}","")).strip() or "?"
        sample_lines.append(f"• {r.get('Emp_Code','')} — Shift: {r.get('Shift','')} | Today: {today_val}")
    msg = (f"📋 Plan Review\n{'─'*28}\n"
           f"Department: {dept}\nMonth: {month}\n"
           f"Employees: {len(plan_rows)}\nStatus: {status}\n\n"
           "Sample (first 5):\n" + "\n".join(sample_lines))
    kb = []
    if role in ("HR_Staff","HR_Manager","Bot_Manager") and status == "Draft":
        kb.append([InlineKeyboardButton("✅ Mark HR Reviewed",
                                        callback_data=f"sched_action_hr_{month.replace('/','_')}_{dept[:10]}")])
    if role in APPROVE_ROLES and status in ("Draft","HR_Review"):
        kb.append([InlineKeyboardButton("✅ Approve Plan",
                                        callback_data=f"sched_action_approve_{month.replace('/','_')}_{dept[:10]}")])
        kb.append([InlineKeyboardButton("❌ Reject Plan",
                                        callback_data=f"sched_action_reject_{month.replace('/','_')}_{dept[:10]}")])
    kb.append([_bsch(), _bm()])
    await q.edit_message_text(msg, reply_markup=InlineKeyboardMarkup(kb))


async def sched_action_cb(update, context):
    q = update.callback_query; await q.answer()
    ec, role, _ = _get_user(str(q.from_user.id))
    data   = q.data.replace("sched_action_","")
    parts  = data.split("_")
    action = parts[0]   # hr / approve / reject
    month  = f"{parts[1]}/{parts[2]}" if len(parts) >= 3 else ""
    dept   = "_".join(parts[3:]) if len(parts) > 3 else ""
    new_status_map = {"hr": "HR_Review", "approve": "Approved", "reject": "Draft"}
    new_status = new_status_map.get(action, "Draft")
    try:
        ws   = get_sheet("Monthly_Schedule")
        rows = ws.get_all_values()
        hdr  = rows[0] if rows else []
        try: month_col = hdr.index("Month") + 1
        except: month_col = 2
        try: dept_col  = hdr.index("Department") + 1
        except: dept_col = 3
        try: status_col = hdr.index("Status") + 1
        except: status_col = 41  # after Plan_ID,Month,Dept,Loc,Emp,Name,Shift,D1-D31 = 7+31=38+2=40, Status=41
        updated = 0
        for i, r in enumerate(rows):
            if i == 0: continue
            if (len(r) >= max(month_col, dept_col)
                    and r[month_col-1].strip() == month
                    and r[dept_col-1].strip().startswith(dept)):
                ws.update_cell(i+1, status_col, new_status)
                updated += 1
        # If approved, also set Active
        if action == "approve" and updated > 0:
            for i, r in enumerate(rows):
                if i == 0: continue
                if (len(r) >= max(month_col, dept_col)
                        and r[month_col-1].strip() == month
                        and r[dept_col-1].strip().startswith(dept)):
                    ws.update_cell(i+1, status_col, "Active")
        label = {"hr":"HR Reviewed","approve":"Approved","reject":"Returned to Draft"}.get(action,action)
        await q.edit_message_text(
            f"✅ {label}!\n{updated} records updated.\nDept: {dept} | Month: {month}",
            reply_markup=InlineKeyboardMarkup([[_bsch(), _bm()]]))
    except Exception as e:
        await q.edit_message_text(f"❌ {e}", reply_markup=InlineKeyboardMarkup([[_bm()]]))


# ══════════════════════════════════════════════════════════════════════════════
#  PDF EXPORT — LANDSCAPE
# ══════════════════════════════════════════════════════════════════════════════
async def sched_pdf_handler(update, context):
    q = update.callback_query; await q.answer()
    ec, role, dept = _get_user(str(q.from_user.id))
    if not _HAS_FPDF:
        await q.edit_message_text("❌ fpdf2 not installed.",
                                  reply_markup=InlineKeyboardMarkup([[_bm()]])); return
    await q.edit_message_text("⏳ Generating schedule PDF...")
    try:
        all_rows = get_sheet("Monthly_Schedule").get_all_records()
    except Exception as e:
        await q.edit_message_text(f"❌ {e}", reply_markup=InlineKeyboardMarkup([[_bm()]])); return
    month = _current_month_str()
    if role in APPROVE_ROLES:
        rows = [r for r in all_rows if str(r.get("Month","")).strip() == month]
    else:
        rows = [r for r in all_rows
                if str(r.get("Month","")).strip() == month
                and str(r.get("Department","")).strip() == dept]
    if not rows:
        await q.edit_message_text(
            f"📄 No schedule data for {month}.",
            reply_markup=InlineKeyboardMarkup([[_bsch(), _bm()]])); return
    n_days = _month_days(month)
    try:
        pdf_bytes = _gen_pdf(rows, month, dept or "All Departments", n_days)
        filename = f"Schedule_{dept or 'All'}_{month.replace('/','_')}.pdf"
        from drive_utils import upload_to_drive
        drive_url = upload_to_drive(pdf_bytes, filename, "schedules")
        if drive_url:
            await q.edit_message_text(
                f"✅ Schedule PDF ready — {dept or 'All'} — {month}",
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton("📄 View PDF", url=drive_url)],
                    [_bsch(), _bm()]]))
        else:
            bio = io.BytesIO(pdf_bytes); bio.name = filename
            await q.message.reply_document(
                document=bio, filename=filename,
                caption=f"📅 Monthly Schedule — {dept or 'All'} — {month}")
            await q.edit_message_text("✅ PDF generated.",
                                      reply_markup=InlineKeyboardMarkup([[_bsch(), _bm()]]))
    except Exception as e:
        await q.edit_message_text(f"❌ PDF error: {e}",
                                  reply_markup=InlineKeyboardMarkup([[_bsch(), _bm()]]))


def _gen_pdf(rows, month, dept, n_days):
    """Generate landscape schedule PDF and return bytes."""
    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=10)
    pdf.add_page()

    # Header
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, "ROIN WORLD FZE", ln=True, align="C")
    pdf.set_font("Helvetica", "B", 13)
    pdf.cell(0, 8, f"{dept} — Monthly Schedule — {month}", ln=True, align="C")
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(0, 5, f"Generated: {datetime.now().strftime('%d/%m/%Y %H:%M')}", ln=True, align="C")
    pdf.ln(3)

    # Calculate col widths
    # Name col=35mm, then n_days cols share remaining width
    page_w   = 277  # A4 landscape inner width (~297-20 margins)
    name_w   = 38
    day_w    = max(4.5, (page_w - name_w) / n_days)
    row_h    = 6

    # Day header row
    pdf.set_font("Helvetica", "B", 7)
    pdf.set_fill_color(50, 80, 130)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(name_w, row_h, "Employee", border=1, fill=True, align="C")
    for d in range(1, n_days + 1):
        pdf.cell(day_w, row_h, str(d), border=1, fill=True, align="C")
    pdf.ln()
    pdf.set_text_color(0, 0, 0)

    # Color map for day values
    colors = {
        "W":   (200, 230, 200),  # green
        "N":   (180, 180, 230),  # blue
        "OFF": (240, 240, 240),  # grey
        "V":   (255, 230, 180),  # orange
        "S":   (255, 180, 180),  # red
        "U":   (255, 200, 200),  # light red
        "B":   (200, 220, 255),  # light blue
        "H":   (220, 255, 220),  # light green
        "R":   (240, 240, 200),  # yellow
    }

    pdf.set_font("Helvetica", "", 7)
    for idx, r in enumerate(rows):
        if idx % 2 == 0:
            pdf.set_fill_color(250, 250, 250)
        else:
            pdf.set_fill_color(245, 245, 255)
        name = str(r.get("Emp_Code",""))
        pdf.cell(name_w, row_h, name[:18], border=1, fill=True)
        for d in range(1, n_days + 1):
            val = str(r.get(f"D{d}","")).strip() or ""
            clr = colors.get(val, (255, 255, 255))
            pdf.set_fill_color(*clr)
            pdf.cell(day_w, row_h, val, border=1, fill=True, align="C")
        pdf.ln()

    # Legend
    pdf.ln(3)
    pdf.set_font("Helvetica", "B", 8)
    pdf.cell(0, 5, "Legend:", ln=True)
    pdf.set_font("Helvetica", "", 7)
    legend_items = [
        ("W=Work (Day)", colors["W"]),
        ("N=Night Shift", colors["N"]),
        ("OFF=Day Off", colors["OFF"]),
        ("V=Vacation", colors["V"]),
        ("S=Sick", colors["S"]),
        ("U=Unpaid", colors["U"]),
        ("B=Business", colors["B"]),
        ("H=Holiday", colors["H"]),
        ("R=Rest", colors["R"]),
    ]
    for label, clr in legend_items:
        pdf.set_fill_color(*clr)
        pdf.cell(30, 5, label, border=1, fill=True)
    pdf.ln()

    # Footer
    pdf.ln(2)
    pdf.set_font("Helvetica", "I", 8)
    pdf.cell(0, 5, "Generated by ROIN WORLD FZE HR System", align="C")

    return pdf.output()


# ══════════════════════════════════════════════════════════════════════════════
#  HANDLER REGISTRATION
# ══════════════════════════════════════════════════════════════════════════════
def get_schedule_handlers():
    """Returns ConversationHandlers."""
    upload_h = ConversationHandler(
        entry_points=[CallbackQueryHandler(sched_upload_start, pattern="^sched_upload$")],
        states={
            SCH_UPLOAD:         [MessageHandler(filters.Document.ALL, sched_upload_doc)],
            SCH_UPLOAD_CONFIRM: [CallbackQueryHandler(sched_upload_confirm_cb,
                                                      pattern="^sched_up_(confirm|cancel)$")],
        },
        fallbacks=[MessageHandler(filters.COMMAND, sched_upload_cancel)],
        per_message=False,
    )
    return [upload_h]


def get_schedule_static_handlers():
    """Returns static CallbackQueryHandlers."""
    return [
        CallbackQueryHandler(sched_menu_handler,        pattern="^menu_schedule$"),
        CallbackQueryHandler(sched_view_handler,        pattern="^sched_view$"),
        CallbackQueryHandler(sched_daily_handler,       pattern="^sched_daily$"),
        CallbackQueryHandler(sched_vs_actual_handler,   pattern="^sched_vs_actual$"),
        CallbackQueryHandler(sched_review_list_handler, pattern="^sched_review_list$"),
        CallbackQueryHandler(sched_review_cb,           pattern="^sched_review_"),
        CallbackQueryHandler(sched_action_cb,           pattern="^sched_action_"),
        CallbackQueryHandler(sched_pdf_handler,         pattern="^sched_pdf$"),
    ]
