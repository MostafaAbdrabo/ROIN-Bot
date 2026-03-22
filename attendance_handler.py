"""
ROIN WORLD FZE — Attendance Handler v4
Uses python-calamine for rock-solid .xls/.xlsx reading (any size).
Supports: .xls, .xlsx, .csv, .zip — send directly, no conversion needed.
Branch auto-detected from filename.
"""

from datetime import datetime, timedelta
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (ConversationHandler, CallbackQueryHandler, MessageHandler, filters)
from config import get_sheet, CURRENT_ATTENDANCE_TAB
import gspread, json, os, io

HEADER_ROW=6; DATA_START_ROW=7; CODE_COL=3; DAY_1_COL=6
RAW_TAB_NAME="Raw_ZKT"; GRACE_MINUTES=15
ST_PRESENT="P"; ST_ABSENT="A"; ST_VACATION="V"; ST_SICK="S"
ST_UNPAID="U"; ST_BUSINESS="B"; ST_OFF="OFF"; ST_HOLIDAY="H"
LEAVE_TYPE_MAP={"Paid":ST_VACATION,"Emergency":ST_VACATION,"Sick":ST_SICK,
                "Unpaid":ST_UNPAID,"Business_Trip":ST_BUSINESS}

ATT_TAB_INPUT=300; ATT_CONFIRM=301; UPLOAD_FILES=310

ATTENDANCE_SHEET_ID="1_GEamKcub5g8zUXryHJ8PVHydA3hKSFV88hUt98fCTw"
ZKT_CODE_HEADERS={"ac-no.","ac-no","acno","ac_no"}
ZKT_DATE_HEADERS={"date"}
ZKT_IN_HEADERS={"clock in","clockin","clock_in"}
ZKT_OUT_HEADERS={"clock out","clockout","clock_out"}

BRANCH_MAP={
    "russian":   "Russian Kitchen",
    "egyptian":  "Egyptian Kitchen",
    "pastry":    "Pastry & Soup Kitchen",
    "soup":      "Pastry & Soup Kitchen",
    "foodtruck": "FoodTruck & Bakery",
    "food_truck":"FoodTruck & Bakery",
    "bakery":    "Bakery & FoodTruck",
    "office":    "Office",
    "warehouse": "Warehouse",
    "transport": "Transportation",
    "steward":   "Steward / Cleaning",
    "cleaning":  "Steward / Cleaning",
    "packaging": "Packaging & Delivery",
    "delivery":  "Packaging & Delivery",
    "supply":    "Purchasing / Supply",
    "purchasing":"Purchasing / Supply",
    "quality":   "Quality Control",
    "safety":    "Safety & OH",
    "translation":"Translation",
    "operations":"Operations",
    "housing":   "Housing",
    "finance":   "Finance",
    "admin":     "Administration",
}

def detect_branch(filename):
    fn=filename.lower()
    for key,name in BRANCH_MAP.items():
        if key in fn: return name
    return "Unknown"

def bm(): return InlineKeyboardButton("↩️ Main Menu",callback_data="back_to_menu")

def _get_gspread_client():
    scopes=["https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive"]
    from google.oauth2.service_account import Credentials
    if os.environ.get("RAILWAY_ENVIRONMENT"):
        creds=Credentials.from_service_account_info(
            json.loads(os.environ["GOOGLE_CREDENTIALS"]),scopes=scopes)
    else:
        creds=Credentials.from_service_account_file("credentials.json",scopes=scopes)
    return gspread.authorize(creds)

def get_attendance_sheet(tab):
    return _get_gspread_client().open_by_key(ATTENDANCE_SHEET_ID).worksheet(tab)

def get_attendance_tab_list():
    return [ws.title for ws in _get_gspread_client().open_by_key(
        ATTENDANCE_SHEET_ID).worksheets()]

def parse_date(s):
    if not s: return None
    s=str(s).strip()
    for fmt in ("%d/%m/%Y","%d-%m-%Y","%Y-%m-%d","%m/%d/%Y"):
        try: return datetime.strptime(s,fmt)
        except: continue
    return None

def parse_time(s):
    if not s: return None
    s=str(s).strip()
    if not s or s in ("0","0.0","None","nan","NaT"): return None
    try: return datetime.strptime(s,"%H:%M")
    except:
        try: return datetime.strptime(s,"%H:%M:%S")
        except: return None

def time_diff_hours(ti,to):
    if not ti or not to: return 0.0
    d=(to-ti).total_seconds()
    if d<0: d+=86400
    return d/3600

def is_friday(d): return d.weekday()==4
def is_2nd_saturday(d): return d.weekday()==5 and 8<=d.day<=14


# ══════════════════════════════════════════════════════════════════
#  FILE PARSING — calamine (primary), csv fallback
# ══════════════════════════════════════════════════════════════════

def _find_cols(headers):
    cc=dc=ic=oc=None
    for i,h in enumerate(headers):
        hl=str(h).strip().lower() if h else ""
        if hl in ZKT_CODE_HEADERS: cc=i
        elif hl in ZKT_DATE_HEADERS: dc=i
        elif hl in ZKT_IN_HEADERS: ic=i
        elif hl in ZKT_OUT_HEADERS: oc=i
    return cc,dc,ic,oc

def _clean_code(v):
    v=str(v).strip()
    if v.endswith(".0"): v=v[:-2]
    return v

def _fmt_date(dv):
    if not dv: return ""
    if hasattr(dv,"strftime"): return dv.strftime("%d/%m/%Y")
    s=str(dv).strip()
    # Handle datetime strings from calamine
    if "T" in s:
        try: return datetime.fromisoformat(s).strftime("%d/%m/%Y")
        except: pass
    return s

def _fmt_time(tv):
    if not tv: return ""
    if hasattr(tv,"strftime"): return tv.strftime("%H:%M")
    s=str(tv).strip()
    if s in ("","0","0.0","None","nan","NaT","false","False"): return ""
    # Handle datetime strings from calamine like "1899-12-31T08:00:00"
    if "T" in s:
        try: return datetime.fromisoformat(s).strftime("%H:%M")
        except: pass
    # Handle timedelta strings like "8:00:00"
    if s.count(":")==2 and len(s)<=8:
        return s[:5]  # "08:00:00" → "08:00"
    return s

def _build_rows(data_rows, headers, filename, branch):
    cc,dc,ic,oc=_find_cols(headers)
    if cc is None or dc is None:
        clean_headers=[str(h).strip() for h in headers if h]
        return [],f"{filename}: can't find AC-No./Date. Headers: {clean_headers[:10]}"
    rows=[]
    for r in data_rows:
        if len(r)<=max(cc,dc): continue
        code=_clean_code(r[cc] if r[cc] else "")
        if not code: continue
        ds=_fmt_date(r[dc])
        if not ds: continue
        cin=_fmt_time(r[ic] if ic is not None and ic<len(r) else None)
        cout=_fmt_time(r[oc] if oc is not None and oc<len(r) else None)
        in_br=branch if cin else ""
        out_br=branch if cout else ""
        rows.append([code,ds,cin,cout,in_br,out_br])
    return rows,None


def _parse_calamine(file_bytes, filename, branch):
    """Parse .xls or .xlsx using python-calamine (Rust-based, handles large files)."""
    try:
        from python_calamine import CalamineWorkbook
    except ImportError:
        return [],f"Missing library: python-calamine. Run: pip3 install python-calamine"
    try:
        import tempfile
        ext = os.path.splitext(filename)[1].lower()
        if not ext: ext = ".xls"
        with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as tmp:
            tmp.write(file_bytes)
            tmp_path = tmp.name
        try:
            wb = CalamineWorkbook.from_path(tmp_path)
            data = wb.get_sheet_by_index(0).to_python()
        finally:
            os.unlink(tmp_path)
    except Exception as e:
        return [],f"{filename}: calamine error ({type(e).__name__}: {e})"
    if len(data)<2: return [],f"{filename}: empty"
    return _build_rows(data[1:], data[0], filename, branch)


def _parse_csv(file_bytes, filename, branch):
    """Parse .csv file."""
    import csv
    try:
        text=file_bytes.decode("utf-8",errors="replace")
        all_rows=list(csv.reader(io.StringIO(text)))
    except Exception as e:
        return [],f"{filename}: csv error ({e})"
    if len(all_rows)<2: return [],f"{filename}: empty"
    return _build_rows(all_rows[1:], all_rows[0], filename, branch)


def _parse_openpyxl(file_bytes, filename, branch):
    """Parse .xlsx with openpyxl. Reads ALL tabs, detects branch from tab name."""
    try:
        from openpyxl import load_workbook
        wb=load_workbook(io.BytesIO(file_bytes),read_only=True,data_only=True)
        all_rows_combined=[]
        tabs_read=0
        for ws in wb.worksheets:
            tab_name=ws.title
            # Detect branch from tab name, fallback to filename, fallback to param
            tab_branch=detect_branch(tab_name)
            if tab_branch=="Unknown":
                tab_branch=detect_branch(filename) if detect_branch(filename)!="Unknown" else (branch or "Unknown")
            all_rows=list(ws.iter_rows(values_only=True))
            if len(all_rows)<2: continue
            rows,err=_build_rows(all_rows[1:], list(all_rows[0]), f"{filename}/{tab_name}", tab_branch)
            if not err and rows:
                all_rows_combined.extend(rows)
                tabs_read+=1
                print(f"[ATT] Tab '{tab_name}' → branch '{tab_branch}': {len(rows)} rows")
        wb.close()
    except Exception as e:
        return [],f"{filename}: xlsx error ({type(e).__name__}: {e})"
    if not all_rows_combined: return [],f"{filename}: no valid data in any tab"
    print(f"[ATT] {filename}: {tabs_read} tabs, {len(all_rows_combined)} total rows")
    return all_rows_combined,None


def parse_zkteco_file(file_bytes, filename="unknown", branch=None):
    """Auto-detect format and parse. Returns (rows, error)."""
    if branch is None: branch=detect_branch(filename)
    fn=filename.lower()
    if fn.endswith(".csv"):
        return _parse_csv(file_bytes, filename, branch)
    if fn.endswith(".xlsx"):
        return _parse_openpyxl(file_bytes, filename, branch)
    # .xls: try calamine first, fallback to openpyxl
    rows,err=_parse_calamine(file_bytes, filename, branch)
    if err:
        rows2,err2=_parse_openpyxl(file_bytes, filename, branch)
        if not err2 and rows2: return rows2,None
        return rows,err
    return rows,None


def write_to_raw_zkt(all_rows,clear_first=True):
    ws=get_attendance_sheet(RAW_TAB_NAME)
    if clear_first:
        ex=ws.get_all_values()
        if len(ex)>1: ws.batch_clear([f"A2:F{len(ex)}"])
    if not all_rows: return 0
    ws.append_rows(all_rows,value_input_option="USER_ENTERED")
    return len(all_rows)


# ══════════════════════════════════════════════════════════════════
#  DATA LOADING
# ══════════════════════════════════════════════════════════════════

def load_holidays():
    h=set()
    try:
        for r in get_sheet("Holidays").get_all_records():
            ds=str(r.get("Date","")).strip()
            if ds:
                d=parse_date(ds)
                if d: h.add(d.date())
    except: pass
    return h

def load_shift_hours():
    s={}
    try:
        for r in get_sheet("Employee_DB").get_all_records():
            ec=str(r.get("Emp_Code","")).strip()
            if ec:
                try: s[ec]=float(r.get("Shift_Hours",9))
                except: s[ec]=9.0
    except: pass
    return s

def load_off_types():
    """Load Off_Type from Employee_DB. Returns {emp_code: 'Rotating'|'Friday_Only'|'Friday_2ndSat'}"""
    ot={}
    try:
        for r in get_sheet("Employee_DB").get_all_records():
            ec=str(r.get("Emp_Code","")).strip()
            if ec: ot[ec]=str(r.get("Off_Type","Friday_2ndSat")).strip()
    except: pass
    return ot

def load_approved_leaves():
    leaves=[]
    try:
        rows=get_sheet("Leave_Log").get_all_values()
        for i,r in enumerate(rows):
            if i==0 or len(r)<16 or r[15].strip()!="Approved": continue
            ec=str(r[1]).strip(); rt=str(r[2]).strip()
            if rt=="Missing_Punch": continue  # handled separately
            sd=parse_date(r[3]); ed=parse_date(r[4])
            if ec and sd and ed:
                leaves.append({"emp_code":ec,"type":rt,"start":sd.date(),"end":ed.date()})
    except: pass
    return leaves

def load_approved_missing_punches():
    """Return set of (emp_code, date_str) for approved Missing_Punch requests."""
    mp=set()
    try:
        rows=get_sheet("Leave_Log").get_all_values()
        for i,r in enumerate(rows):
            if i==0 or len(r)<16 or r[15].strip()!="Approved": continue
            if str(r[2]).strip()!="Missing_Punch": continue
            ec=str(r[1]).strip(); sd=parse_date(r[3])
            if ec and sd: mp.add((ec,sd.strftime("%Y-%m-%d")))
    except: pass
    return mp

def find_leave_for(leaves,ec,d):
    dd=d.date() if hasattr(d,"date") else d
    for lv in leaves:
        if lv["emp_code"]==str(ec).strip() and lv["start"]<=dd<=lv["end"]:
            return lv["type"]
    return None

def has_missing_punch_approval(mp_set,ec,d):
    key=(str(ec).strip(),d.strftime("%Y-%m-%d") if hasattr(d,"strftime") else str(d))
    return key in mp_set

def load_raw_zkt():
    ws=get_attendance_sheet(RAW_TAB_NAME); all_rows=ws.get_all_values()
    punches={}; skipped=0
    for i,row in enumerate(all_rows):
        if i==0 or len(row)<2: continue
        code=str(row[0]).strip(); ds=str(row[1]).strip()
        cin=str(row[2]).strip() if len(row)>2 else ""
        cout=str(row[3]).strip() if len(row)>3 else ""
        in_br=str(row[4]).strip() if len(row)>4 else ""
        out_br=str(row[5]).strip() if len(row)>5 else ""
        if not code or not ds: skipped+=1; continue
        d=parse_date(ds)
        if not d: skipped+=1; continue
        key=(code,d.strftime("%Y-%m-%d"))
        if key not in punches: punches[key]={"ins":[],"outs":[],"date":d}
        if cin: punches[key]["ins"].append((cin,in_br))
        if cout: punches[key]["outs"].append((cout,out_br))
    print(f"[ATT] Raw_ZKT: {len(all_rows)-1} rows, {len(punches)} combos, {skipped} skipped")
    return punches

def best_punch(pd):
    ins=[(parse_time(t),b) for t,b in pd.get("ins",[]) if parse_time(t)]
    outs=[(parse_time(t),b) for t,b in pd.get("outs",[]) if parse_time(t)]
    if ins: t_in,in_br=min(ins,key=lambda x:x[0])
    else: t_in=None; in_br=""
    if outs: t_out,out_br=max(outs,key=lambda x:x[0])
    else: t_out=None; out_br=""
    return t_in,t_out,in_br,out_br


# ══════════════════════════════════════════════════════════════════
#  MAIN PROCESSING
# ══════════════════════════════════════════════════════════════════

def process_attendance(target_tab_name):
    print(f"[ATT] Processing: {target_tab_name}")
    punches=load_raw_zkt()
    tw=get_attendance_sheet(target_tab_name); td=tw.get_all_values()
    leaves=load_approved_leaves(); holidays=load_holidays()
    shift_hours=load_shift_hours(); off_types=load_off_types()
    missing_punches=load_approved_missing_punches()

    code_to_row={}
    for i,row in enumerate(td):
        if i<DATA_START_ROW-1: continue
        if len(row)>=CODE_COL:
            ec=str(row[CODE_COL-1]).strip()
            if ec and ec not in ("Code","code",""): code_to_row[ec]=i+1

    all_dates=sorted(set(p["date"] for p in punches.values()))
    if not all_dates: return {"error":"No valid dates in Raw_ZKT."}

    raw_codes=set(k[0] for k in punches.keys())
    updates=[]; stats={}

    for d in all_dates:
        col=DAY_1_COL+d.day-1; dd=d.date()
        ds=d.strftime("%d/%m/%Y"); stats[ds]={}
        for ec,sr in code_to_row.items():
            status=None

            # Step 1: Approved leave always wins
            lt=find_leave_for(leaves,ec,d)
            if lt:
                status=LEAVE_TYPE_MAP.get(lt,ST_VACATION)
                updates.append((sr,col,status))
                stats[ds][status]=stats[ds].get(status,0)+1
                continue

            # Step 2: Check punch data — PUNCH ALWAYS WINS over schedule
            key=(ec,d.strftime("%Y-%m-%d")); punch=punches.get(key)
            if punch:
                ti,to,in_br,out_br=best_punch(punch)
                sh=shift_hours.get(ec,9.0); gr=GRACE_MINUTES/60.0
                if ti and to:
                    diff=time_diff_hours(ti,to)
                    status=ST_PRESENT if diff>=(sh-gr) else ST_ABSENT
                elif ti or to:
                    # Has partial punch (in or out only)
                    # Check for approved Missing_Punch request → counts as Present
                    if has_missing_punch_approval(missing_punches,ec,dd):
                        status=ST_PRESENT
                    else:
                        status=ST_ABSENT
                else:
                    # In ZKTeco but no actual punch times
                    # Check if it's their off day
                    ot=off_types.get(ec,"Friday_2ndSat")
                    if dd in holidays:
                        status=ST_HOLIDAY
                    elif ot=="Rotating":
                        status=ST_OFF
                    elif ot=="Friday" and is_friday(d):
                        status=ST_OFF
                    elif ot=="Friday_2ndSat" and (is_friday(d) or is_2nd_saturday(d)):
                        status=ST_OFF
                    else:
                        status=ST_ABSENT
            else:
                # Not in ZKTeco at all — skip (leave cell empty)
                continue

            if status:
                updates.append((sr,col,status))
                stats[ds][status]=stats[ds].get(status,0)+1

    print(f"[ATT] Writing {len(updates)} cells...")
    if updates:
        cells=[{"range":f"'{target_tab_name}'!{gspread.utils.rowcol_to_a1(r,c)}",
                "values":[[v]]} for r,c,v in updates]
        for i in range(0,len(cells),500):
            tw.spreadsheet.values_batch_update(
                body={"valueInputOption":"USER_ENTERED","data":cells[i:i+500]})

    return {"tab":target_tab_name,"dates":[d.strftime("%d/%m/%Y") for d in all_dates],
            "employees_in_tab":len(code_to_row),"raw_records":len(punches),
            "raw_employees":len(raw_codes),
            "matched":len(raw_codes.intersection(code_to_row.keys())),
            "unmatched_codes":sorted(raw_codes-code_to_row.keys())[:20],
            "unmatched_count":len(raw_codes-code_to_row.keys()),
            "cells_written":len(updates),"stats":stats}


# ══════════════════════════════════════════════════════════════════
#  MY ATTENDANCE SUMMARY (personal view for employees)
# ══════════════════════════════════════════════════════════════════

def get_my_attendance_summary(emp_code, tab_name):
    """Return attendance counts + day-by-day list for emp_code in tab, up to today."""
    ws = get_attendance_sheet(tab_name)
    data = ws.get_all_values()
    emp_row = None
    for i, row in enumerate(data):
        if i < DATA_START_ROW - 1: continue
        if len(row) >= CODE_COL and str(row[CODE_COL - 1]).strip() == str(emp_code).strip():
            emp_row = row; break
    if emp_row is None:
        return None
    today_day = datetime.now().day
    counts = {ST_PRESENT: 0, ST_ABSENT: 0, ST_VACATION: 0, ST_SICK: 0,
              ST_UNPAID: 0, ST_BUSINESS: 0, ST_OFF: 0, ST_HOLIDAY: 0}
    days = []
    for day in range(1, today_day + 1):
        col_idx = DAY_1_COL - 1 + (day - 1)
        if col_idx >= len(emp_row): break
        val = str(emp_row[col_idx]).strip().upper()
        if val == ST_PRESENT: counts[ST_PRESENT] += 1
        elif val == ST_ABSENT: counts[ST_ABSENT] += 1
        elif val == ST_VACATION: counts[ST_VACATION] += 1
        elif val == ST_SICK: counts[ST_SICK] += 1
        elif val == ST_UNPAID: counts[ST_UNPAID] += 1
        elif val == ST_BUSINESS: counts[ST_BUSINESS] += 1
        elif val == "OFF": counts[ST_OFF] += 1
        elif val == ST_HOLIDAY: counts[ST_HOLIDAY] += 1
        days.append((day, val if val else "—"))
    leave = counts[ST_VACATION] + counts[ST_SICK] + counts[ST_UNPAID] + counts[ST_BUSINESS]
    working = counts[ST_PRESENT] + counts[ST_ABSENT] + leave
    rate = round(counts[ST_PRESENT] / working * 100) if working > 0 else 0
    return {"present": counts[ST_PRESENT], "absent": counts[ST_ABSENT],
            "leave": leave, "off": counts[ST_OFF], "holiday": counts[ST_HOLIDAY], "rate": rate,
            "days": days}


# ══════════════════════════════════════════════════════════════════
#  ATTENDANCE NAVIGATION — year → month → summary
# ══════════════════════════════════════════════════════════════════

MONTHS_SHORT = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
MONTHS_FULL  = ["January","February","March","April","May","June",
                "July","August","September","October","November","December"]

def _att_years():
    y = datetime.now().year; return [y, y - 1, y - 2]

def _get_role_for_att(tid):
    for i, r in enumerate(get_sheet("User_Registry").get_all_values()):
        if i == 0: continue
        if r[1].strip() == str(tid):
            return r[0].strip(), (r[3].strip() if len(r) > 3 else "Employee")
    return None, "Employee"

def _get_team_for_att(my_code, my_role):
    team = []
    for r in get_sheet("Employee_DB").get_all_records():
        ec = str(r.get("Emp_Code", "")).strip()
        name = r.get("Full_Name", ec)
        if my_role == "Supervisor" and str(r.get("Supervisor_Code", "")).strip() == str(my_code):
            team.append((ec, name))
        elif my_role in ("Direct_Manager", "HR_Manager") and str(r.get("Manager_Code", "")).strip() == str(my_code):
            team.append((ec, name))
    return team

def _get_emp_name(ec):
    try:
        for r in get_sheet("Employee_DB").get_all_records():
            if str(r.get("Emp_Code", "")).strip() == str(ec): return r.get("Full_Name", ec)
    except: pass
    return str(ec)

def _year_kb(next_prefix, back_cb):
    return InlineKeyboardMarkup([
        [InlineKeyboardButton(str(y), callback_data=f"{next_prefix}{y}") for y in _att_years()],
        [InlineKeyboardButton("↩️ Back", callback_data=back_cb), bm()]])

def _month_kb(next_prefix, back_cb):
    rows = [[InlineKeyboardButton(MONTHS_SHORT[j], callback_data=f"{next_prefix}{j+1:02d}")
             for j in range(i, min(i + 3, 12))] for i in range(0, 12, 3)]
    rows.append([InlineKeyboardButton("↩️ Back", callback_data=back_cb), bm()])
    return InlineKeyboardMarkup(rows)

async def _show_att_summary(q, ec, year, month, back_cb):
    mo_name = MONTHS_FULL[int(month) - 1]
    name = _get_emp_name(ec) if ec else "?"
    summary = None
    now = datetime.now()
    if int(year) == now.year and int(month) == now.month:
        try: summary = get_my_attendance_summary(ec, CURRENT_ATTENDANCE_TAB)
        except: pass
    if summary:
        p, a, lv, o, h = summary["present"], summary["absent"], summary["leave"], summary["off"], summary["holiday"]
        rate = f"{summary['rate']}%"; note = ""
        # Build compact day-by-day
        chunks = [f"{d}:{s}" for d, s in summary.get("days", [])]
        lines = []; cur = ""
        for ch in chunks:
            if len(cur) + len(ch) + 1 > 52:
                lines.append(cur.strip()); cur = ch
            else:
                cur += (" " if cur else "") + ch
        if cur: lines.append(cur.strip())
        day_block = "\n".join(lines)
    else:
        p = a = lv = o = h = "—"; rate = "—"
        note = "\n\n⚠️ Live data available for the current configured month only."
        day_block = ""
    msg = (f"🕐 {name}\n{mo_name} {year}\n{'─' * 28}\n\n"
           f"✅ Present:         {p}\n"
           f"❌ Absent:          {a}\n"
           f"🏖 Leave:           {lv}\n"
           f"📴 Off Days:        {o}\n"
           f"🎌 Holidays:        {h}\n"
           f"{'─' * 28}\n"
           f"📈 Attendance Rate: {rate}{note}")
    if day_block:
        msg += f"\n\n📅 Day-by-Day:\n{day_block}"
    await q.edit_message_text(msg, reply_markup=InlineKeyboardMarkup(
        [[InlineKeyboardButton("↩️ Back", callback_data=back_cb), bm()]]))

# ─── Entry handler (replaces old my_attendance_handler in bot.py) ──────────

async def my_attendance_handler(update, context):
    q = update.callback_query; await q.answer()
    _, my_role = _get_role_for_att(str(q.from_user.id))
    kb = [[InlineKeyboardButton("👤 My Attendance", callback_data="att_own")]]
    if my_role in ("Supervisor", "Direct_Manager"):
        kb.append([InlineKeyboardButton("👥 My Team's Attendance", callback_data="att_team")])
    elif my_role in ("HR_Staff", "HR_Manager", "Director"):
        kb.append([InlineKeyboardButton("👥 Team Attendance", callback_data="att_team")])
        kb.append([InlineKeyboardButton("🏢 Company Overview", callback_data="att_comp")])
    kb.append([bm()])
    await q.edit_message_text("🕐 Attendance\n\nSelect an option:", reply_markup=InlineKeyboardMarkup(kb))

# ─── Own attendance ─────────────────────────────────────────────────────────

async def att_own_handler(update, context):
    q = update.callback_query; await q.answer()
    await q.edit_message_text("🕐 My Attendance\n\nSelect year:",
                               reply_markup=_year_kb("att_oyr_", "menu_my_attendance"))

async def att_own_year_handler(update, context):
    q = update.callback_query; await q.answer()
    year = q.data.replace("att_oyr_", "")
    await q.edit_message_text(f"🕐 My Attendance — {year}\n\nSelect month:",
                               reply_markup=_month_kb(f"att_omo_{year}_", "att_own"))

async def att_own_month_handler(update, context):
    q = update.callback_query; await q.answer()
    year, month = q.data.replace("att_omo_", "").rsplit("_", 1)
    ec, _ = _get_role_for_att(str(q.from_user.id))
    await _show_att_summary(q, ec, year, month, f"att_oyr_{year}")

# ─── Team attendance ────────────────────────────────────────────────────────

async def att_team_handler(update, context):
    q = update.callback_query; await q.answer()
    await q.edit_message_text("⏳ Loading team...")
    my_code, my_role = _get_role_for_att(str(q.from_user.id))
    team = _get_team_for_att(my_code, my_role)
    if not team:
        await q.edit_message_text("ℹ️ No team members found.",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("↩️ Back", callback_data="menu_my_attendance"), bm()]])); return
    kb = [[InlineKeyboardButton(f"👤 {name} ({ec})", callback_data=f"att_te_{ec}")] for ec, name in team[:20]]
    kb.append([InlineKeyboardButton("↩️ Back", callback_data="menu_my_attendance"), bm()])
    await q.edit_message_text("👥 Team Attendance\n\nSelect employee:", reply_markup=InlineKeyboardMarkup(kb))

async def att_team_emp_handler(update, context):
    q = update.callback_query; await q.answer()
    ec = q.data.replace("att_te_", "")
    await q.edit_message_text(f"🕐 {_get_emp_name(ec)}\n\nSelect year:",
                               reply_markup=_year_kb(f"att_tyr_{ec}_", "att_team"))

async def att_team_year_handler(update, context):
    q = update.callback_query; await q.answer()
    ec, year = q.data.replace("att_tyr_", "").rsplit("_", 1)
    await q.edit_message_text(f"🕐 {_get_emp_name(ec)} — {year}\n\nSelect month:",
                               reply_markup=_month_kb(f"att_tmo_{ec}_{year}_", f"att_te_{ec}"))

async def att_team_month_handler(update, context):
    q = update.callback_query; await q.answer()
    parts = q.data.replace("att_tmo_", "").rsplit("_", 2)
    ec, year, month = parts[0], parts[1], parts[2]
    await _show_att_summary(q, ec, year, month, f"att_tyr_{ec}_{year}")

# ─── Company attendance (HR / Director) ─────────────────────────────────────

async def att_comp_handler(update, context):
    q = update.callback_query; await q.answer()
    await q.edit_message_text("⏳ Loading...")
    try:
        depts = sorted(set(str(r.get("Department", "")).strip()
                           for r in get_sheet("Employee_DB").get_all_records()
                           if str(r.get("Department", "")).strip()))
        kb = [[InlineKeyboardButton(f"🏢 {d}", callback_data=f"att_cd_{d}")] for d in depts]
        kb.append([InlineKeyboardButton("↩️ Back", callback_data="menu_my_attendance"), bm()])
        await q.edit_message_text("🏢 Company Attendance\n\nSelect department:",
                                   reply_markup=InlineKeyboardMarkup(kb))
    except Exception as e:
        await q.edit_message_text(f"❌ {e}", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("↩️ Back", callback_data="menu_my_attendance"), bm()]]))

async def att_comp_dept_handler(update, context):
    q = update.callback_query; await q.answer()
    dept = q.data.replace("att_cd_", "")
    await q.edit_message_text("⏳ Loading...")
    try:
        emps = [(str(r.get("Emp_Code", "")), r.get("Full_Name", "?"))
                for r in get_sheet("Employee_DB").get_all_records()
                if str(r.get("Department", "")).strip() == dept and str(r.get("Emp_Code", "")).strip()]
        if not emps:
            await q.edit_message_text(f"ℹ️ No employees in {dept}.",
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("↩️ Back", callback_data="att_comp"), bm()]])); return
        kb = [[InlineKeyboardButton(f"👤 {nm} ({ec})", callback_data=f"att_ce_{ec}")] for ec, nm in emps[:25]]
        kb.append([InlineKeyboardButton("↩️ Back", callback_data="att_comp"), bm()])
        await q.edit_message_text(f"🏢 {dept}\n\nSelect employee:",
                                   reply_markup=InlineKeyboardMarkup(kb))
    except Exception as e:
        await q.edit_message_text(f"❌ {e}", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("↩️ Back", callback_data="att_comp"), bm()]]))

async def att_comp_emp_handler(update, context):
    q = update.callback_query; await q.answer()
    ec = q.data.replace("att_ce_", "")
    await q.edit_message_text(f"🕐 {_get_emp_name(ec)}\n\nSelect year:",
                               reply_markup=_year_kb(f"att_cyr_{ec}_", "att_comp"))

async def att_comp_year_handler(update, context):
    q = update.callback_query; await q.answer()
    ec, year = q.data.replace("att_cyr_", "").rsplit("_", 1)
    await q.edit_message_text(f"🕐 {_get_emp_name(ec)} — {year}\n\nSelect month:",
                               reply_markup=_month_kb(f"att_cmo_{ec}_{year}_", f"att_ce_{ec}"))

async def att_comp_month_handler(update, context):
    q = update.callback_query; await q.answer()
    parts = q.data.replace("att_cmo_", "").rsplit("_", 2)
    ec, year, month = parts[0], parts[1], parts[2]
    await _show_att_summary(q, ec, year, month, f"att_cyr_{ec}_{year}")


# ══════════════════════════════════════════════════════════════════
#  BOT — ATTENDANCE MENU
# ══════════════════════════════════════════════════════════════════

async def attendance_menu(update,context):
    q=update.callback_query; await q.answer()
    kb=[[InlineKeyboardButton("📥 Upload ZKTeco Files",callback_data="att_upload")],
        [InlineKeyboardButton("📤 Process Attendance",callback_data="att_process")],
        [InlineKeyboardButton("📊 Monthly Summary",callback_data="att_summary")],
        [InlineKeyboardButton("↩️ Back",callback_data="menu_hr_tools"),bm()]]
    await q.edit_message_text(
        "🕐 *Attendance Overview*\n\n"
        "• *Upload ZKTeco Files* — send .xls/.xlsx/.csv/.zip\n"
        "• *Process Attendance* — calculate P/A/V/S/U/B/OFF/H\n"
        "• *Monthly Summary* — coming soon\n",
        parse_mode="Markdown",reply_markup=InlineKeyboardMarkup(kb))

async def att_summary_placeholder(update,context):
    q=update.callback_query; await q.answer()
    await q.edit_message_text("📊 Monthly Summary\n\nComing soon.",
        reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton(
            "↩️ Back",callback_data="menu_attendance_overview"),bm()]]))


# ══════════════════════════════════════════════════════════════════
#  UPLOAD FLOW
# ══════════════════════════════════════════════════════════════════

async def upload_start(update,context):
    q=update.callback_query; await q.answer()
    context.user_data["zkt_rows"]=[]; context.user_data["zkt_files"]=0
    context.user_data["zkt_errors"]=[]
    kb=[[InlineKeyboardButton("✅ Done — Write to Sheet",callback_data="upload_done")],
        [InlineKeyboardButton("❌ Cancel",callback_data="upload_cancel")],[bm()]]
    await q.edit_message_text(
        "📥 *Upload ZKTeco Files*\n\n"
        "Send your .xls files directly — any size.\n"
        "Also supports .xlsx, .csv, and .zip\n"
        "Branch auto-detected from filename.\n\n"
        "Tap *Done* when all files are sent.",
        parse_mode="Markdown",reply_markup=InlineKeyboardMarkup(kb))
    return UPLOAD_FILES

async def upload_receive_file(update,context):
    doc=update.message.document
    if not doc:
        await update.message.reply_text("⚠️ Send a file (.xls, .xlsx, .csv, or .zip)")
        return UPLOAD_FILES
    fn=doc.file_name or "unknown"
    await update.message.reply_text(f"⏳ Processing {fn}...")
    try:
        tf=await context.bot.get_file(doc.file_id); buf=io.BytesIO()
        await tf.download_to_memory(buf)
        file_bytes=buf.getvalue()

        import zipfile
        all_rows=[]; all_errors=[]; files_count=0

        if fn.lower().endswith(".zip"):
            with zipfile.ZipFile(io.BytesIO(file_bytes)) as zf:
                valid=[n for n in zf.namelist()
                       if not n.startswith("__MACOSX") and not n.startswith(".")
                       and (n.lower().endswith(".xls") or n.lower().endswith(".xlsx")
                            or n.lower().endswith(".csv"))]
                if not valid:
                    kb=[[InlineKeyboardButton("✅ Done",callback_data="upload_done")],
                        [InlineKeyboardButton("❌ Cancel",callback_data="upload_cancel")],[bm()]]
                    await update.message.reply_text(f"⚠️ No data files in {fn}",
                        reply_markup=InlineKeyboardMarkup(kb))
                    return UPLOAD_FILES
                for name in valid:
                    fbytes=zf.read(name)
                    rows,error=parse_zkteco_file(fbytes,name)
                    if error: all_errors.append(error)
                    else: all_rows.extend(rows); files_count+=1
        else:
            rows,error=parse_zkteco_file(file_bytes,fn)
            if error: all_errors.append(error)
            else: all_rows=rows; files_count=1

        if all_errors and not all_rows:
            for e in all_errors: context.user_data["zkt_errors"].append(e)
            kb=[[InlineKeyboardButton("✅ Done",callback_data="upload_done")],
                [InlineKeyboardButton("❌ Cancel",callback_data="upload_cancel")],[bm()]]
            await update.message.reply_text(
                "⚠️ Errors:\n"+"\n".join(f"• {e}" for e in all_errors)+
                "\n\nSend another or tap Done.",
                reply_markup=InlineKeyboardMarkup(kb))
            return UPLOAD_FILES

        if all_errors:
            for e in all_errors: context.user_data["zkt_errors"].append(e)

        context.user_data["zkt_rows"].extend(all_rows)
        context.user_data["zkt_files"]+=files_count
        total=len(context.user_data["zkt_rows"])
        files=context.user_data["zkt_files"]
        dates=set(); branches=set()
        for r in all_rows:
            d=parse_date(r[1])
            if d: dates.add(d.strftime("%d/%m/%Y"))
            if len(r)>4 and r[4]: branches.add(r[4])
            if len(r)>5 and r[5]: branches.add(r[5])
        kb=[[InlineKeyboardButton("✅ Done — Write to Sheet",callback_data="upload_done")],
            [InlineKeyboardButton("❌ Cancel",callback_data="upload_cancel")],[bm()]]
        zip_note=f" ({files_count} files from zip)" if fn.lower().endswith(".zip") else ""
        br_note=f"\nBranch: {', '.join(sorted(branches))}" if branches else ""
        err_note=f"\n⚠️ {len(all_errors)} file(s) had errors" if all_errors else ""
        await update.message.reply_text(
            f"✅ *{fn}*{zip_note}\nRecords: {len(all_rows)}\n"
            f"Dates: {', '.join(sorted(dates))}{br_note}"
            f"{err_note}\n\n📊 Total: {files} files, {total} records\n\nSend more or tap *Done*.",
            parse_mode="Markdown",reply_markup=InlineKeyboardMarkup(kb))
    except Exception as e:
        import traceback; traceback.print_exc()
        err_msg=f"{type(e).__name__}: {e}" if str(e) else f"{type(e).__name__}"
        context.user_data["zkt_errors"].append(f"{fn}: {err_msg}")
        kb=[[InlineKeyboardButton("✅ Done",callback_data="upload_done")],
            [InlineKeyboardButton("❌ Cancel",callback_data="upload_cancel")],[bm()]]
        await update.message.reply_text(f"⚠️ Error: {err_msg}\n\nSend another or tap Done.",
            reply_markup=InlineKeyboardMarkup(kb))
    return UPLOAD_FILES

async def upload_done(update,context):
    q=update.callback_query; await q.answer()
    rows=context.user_data.get("zkt_rows",[])
    files=context.user_data.get("zkt_files",0)
    errors=context.user_data.get("zkt_errors",[])
    if not rows:
        kb=[[InlineKeyboardButton("↩️ Back",callback_data="menu_attendance_overview"),bm()]]
        msg="❌ No data."
        if errors: msg+="\n\n"+"\n".join(f"• {e}" for e in errors)
        await q.edit_message_text(msg,reply_markup=InlineKeyboardMarkup(kb))
        return ConversationHandler.END
    await q.edit_message_text(f"⏳ Writing {len(rows)} records to Raw_ZKT...")
    try:
        count=write_to_raw_zkt(rows,clear_first=True)
        dates=set(); codes=set()
        for r in rows:
            codes.add(r[0]); d=parse_date(r[1])
            if d: dates.add(d.strftime("%d/%m/%Y"))
        msg=(f"✅ *Upload Complete!*\n\nFiles: {files}\nRecords: {count}\n"
             f"Employees: {len(codes)}\nDates: {', '.join(sorted(dates))}\n")
        if errors: msg+=f"\n⚠️ Errors:\n"+"\n".join(f"• {e}" for e in errors)
        msg+="\n\nNow tap *Process Attendance* to calculate."
        kb=[[InlineKeyboardButton("📤 Process Attendance",callback_data="att_process")],
            [InlineKeyboardButton("↩️ Back",callback_data="menu_attendance_overview"),bm()]]
        await q.edit_message_text(msg,parse_mode="Markdown",reply_markup=InlineKeyboardMarkup(kb))
    except Exception as e:
        kb=[[InlineKeyboardButton("↩️ Back",callback_data="menu_attendance_overview"),bm()]]
        await q.edit_message_text(f"❌ Write failed: {e}",reply_markup=InlineKeyboardMarkup(kb))
    return ConversationHandler.END

async def upload_cancel(update,context):
    q=update.callback_query; await q.answer()
    context.user_data.pop("zkt_rows",None)
    await attendance_menu(update,context)
    return ConversationHandler.END

async def upload_back_menu(update,context):
    q=update.callback_query; await q.answer()
    from approval_handler import back_to_menu_handler
    await back_to_menu_handler(update,context)
    return ConversationHandler.END


# ══════════════════════════════════════════════════════════════════
#  PROCESS FLOW
# ══════════════════════════════════════════════════════════════════

async def att_start_process(update,context):
    q=update.callback_query; await q.answer()
    try:
        tabs=get_attendance_tab_list()
        tl="\n".join(f"  • {t}" for t in tabs if t!=RAW_TAB_NAME)
    except: tl="  (could not load)"
    kb=[[InlineKeyboardButton("↩️ Back",callback_data="menu_attendance_overview"),bm()]]
    await q.edit_message_text(
        f"📤 *Process Attendance*\n\nTabs:\n{tl}\n\nType exact tab name:",
        parse_mode="Markdown",reply_markup=InlineKeyboardMarkup(kb))
    return ATT_TAB_INPUT

async def att_receive_tab(update,context):
    tn=update.message.text.strip(); context.user_data["att_target_tab"]=tn
    await update.message.reply_text("⏳ Reading...")
    try:
        try:
            tw=get_attendance_sheet(tn); td=tw.get_all_values()
            ec=sum(1 for i,r in enumerate(td) if i>=DATA_START_ROW-1
                   and len(r)>=CODE_COL and str(r[CODE_COL-1]).strip())
        except gspread.exceptions.WorksheetNotFound:
            kb=[[InlineKeyboardButton("↩️ Back",callback_data="menu_attendance_overview"),bm()]]
            await update.message.reply_text(f"❌ Tab '{tn}' not found.",
                reply_markup=InlineKeyboardMarkup(kb))
            return ConversationHandler.END
        try:
            rw=get_attendance_sheet(RAW_TAB_NAME); rd=rw.get_all_values(); rc=len(rd)-1
        except gspread.exceptions.WorksheetNotFound:
            kb=[[InlineKeyboardButton("↩️ Back",callback_data="menu_attendance_overview"),bm()]]
            await update.message.reply_text(
                f"❌ '{RAW_TAB_NAME}' not found. Upload files first.",
                reply_markup=InlineKeyboardMarkup(kb))
            return ConversationHandler.END
        if rc<=0:
            kb=[[InlineKeyboardButton("📥 Upload",callback_data="att_upload"),
                 InlineKeyboardButton("↩️ Back",callback_data="menu_attendance_overview")],[bm()]]
            await update.message.reply_text("❌ Raw_ZKT empty.",
                reply_markup=InlineKeyboardMarkup(kb))
            return ConversationHandler.END
        dates=set()
        for i,row in enumerate(rd):
            if i==0: continue
            if len(row)>1:
                d=parse_date(str(row[1]))
                if d: dates.add(d.strftime("%d/%m/%Y"))
        kb=[[InlineKeyboardButton("✅ Process Now",callback_data="att_execute")],
            [InlineKeyboardButton("❌ Cancel",callback_data="att_cancel")],[bm()]]
        dd=", ".join(sorted(dates)) if len(dates)<=10 else f"{len(dates)} dates"
        await update.message.reply_text(
            f"📊 *Preview*\n\nTab: `{tn}`\nEmployees: {ec}\nRecords: {rc}\n"
            f"Dates: {dd}\n\n⚠️ Values overwritten.\n\nReady?",
            parse_mode="Markdown",reply_markup=InlineKeyboardMarkup(kb))
        return ATT_CONFIRM
    except Exception as e:
        kb=[[InlineKeyboardButton("↩️ Back",callback_data="menu_attendance_overview"),bm()]]
        await update.message.reply_text(f"❌ {e}",reply_markup=InlineKeyboardMarkup(kb))
        return ConversationHandler.END

async def att_execute(update,context):
    q=update.callback_query; await q.answer()
    tab=context.user_data.get("att_target_tab","")
    if not tab:
        await q.edit_message_text("❌",reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("↩️ Back",callback_data="menu_attendance_overview"),bm()]]))
        return ConversationHandler.END
    await q.edit_message_text("⏳ Processing... 30–60 seconds.")
    try:
        result=process_attendance(tab)
        if "error" in result:
            kb=[[InlineKeyboardButton("↩️ Back",callback_data="menu_attendance_overview"),bm()]]
            await q.edit_message_text(f"❌ {result['error']}",
                reply_markup=InlineKeyboardMarkup(kb))
            return ConversationHandler.END
        lines=[f"✅ *Attendance Processed!*\n",f"Tab: `{result['tab']}`",
               f"Cells written: {result['cells_written']}",
               f"Employees: {result['employees_in_tab']} in tab, "
               f"{result['raw_employees']} in ZKTeco",""]
        for ds,st in result["stats"].items():
            parts=[f"{l}:{st.get(c,0)}" for c,l in
                   [(ST_PRESENT,"P"),(ST_ABSENT,"A"),(ST_OFF,"OFF"),(ST_HOLIDAY,"H"),
                    (ST_VACATION,"V"),(ST_SICK,"S"),(ST_UNPAID,"U"),(ST_BUSINESS,"B")]
                   if st.get(c,0)>0]
            lines.append(f"📅 {ds}: {' | '.join(parts)}")
        if result["unmatched_count"]>0:
            lines.append(f"\n⚠️ {result['unmatched_count']} codes not in tab")
            if result["unmatched_codes"]:
                lines.append(f"  {', '.join(result['unmatched_codes'][:10])}")
        kb=[[InlineKeyboardButton("↩️ Back",callback_data="menu_attendance_overview"),bm()]]
        await q.edit_message_text("\n".join(lines),parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup(kb))
    except Exception as e:
        import traceback; traceback.print_exc()
        kb=[[InlineKeyboardButton("↩️ Back",callback_data="menu_attendance_overview"),bm()]]
        await q.edit_message_text(f"❌ {e}",reply_markup=InlineKeyboardMarkup(kb))
    return ConversationHandler.END

async def att_cancel(update,context):
    q=update.callback_query; await q.answer()
    await attendance_menu(update,context)
    return ConversationHandler.END

async def att_back_menu(update,context):
    q=update.callback_query; await q.answer()
    from approval_handler import back_to_menu_handler
    await back_to_menu_handler(update,context)
    return ConversationHandler.END

async def att_back_to_att(update,context):
    q=update.callback_query; await q.answer()
    await attendance_menu(update,context)
    return ConversationHandler.END


# ══════════════════════════════════════════════════════════════════
#  CONVERSATION HANDLERS
# ══════════════════════════════════════════════════════════════════

def get_upload_conversation_handler():
    return ConversationHandler(
        entry_points=[CallbackQueryHandler(upload_start,pattern="^att_upload$")],
        states={UPLOAD_FILES:[
            MessageHandler(filters.Document.ALL,upload_receive_file),
            CallbackQueryHandler(upload_done,pattern="^upload_done$"),
            CallbackQueryHandler(upload_cancel,pattern="^upload_cancel$"),
            CallbackQueryHandler(upload_back_menu,pattern="^back_to_menu$")]},
        fallbacks=[CallbackQueryHandler(upload_back_menu,pattern="^back_to_menu$"),
                   CallbackQueryHandler(upload_cancel,pattern="^upload_cancel$")],
        per_message=False)

def get_attendance_conversation_handler():
    return ConversationHandler(
        entry_points=[CallbackQueryHandler(att_start_process,pattern="^att_process$")],
        states={
            ATT_TAB_INPUT:[
                MessageHandler(filters.TEXT & ~filters.COMMAND,att_receive_tab),
                CallbackQueryHandler(att_back_to_att,pattern="^menu_attendance_overview$"),
                CallbackQueryHandler(att_back_menu,pattern="^back_to_menu$")],
            ATT_CONFIRM:[
                CallbackQueryHandler(att_execute,pattern="^att_execute$"),
                CallbackQueryHandler(att_cancel,pattern="^att_cancel$"),
                CallbackQueryHandler(att_back_menu,pattern="^back_to_menu$")]},
        fallbacks=[CallbackQueryHandler(att_back_menu,pattern="^back_to_menu$"),
                   CallbackQueryHandler(att_back_to_att,pattern="^menu_attendance_overview$")],
        per_message=False)
